# Import pandas 
import chardet
import pandas as pd
from pandas import *
import openpyxl as xl

def paavam(a):
    wb2 = xl.load_workbook("book1.xlsx")
    ws2 = wb2.worksheets[0]
    ws3=wb2.worksheets[1]



    with open(a, 'rb') as f:
        result = chardet.detect(f.read())  # or readline if the file is large
    d=pd.read_csv(a, encoding=result['encoding'])
    GFG = pd.ExcelWriter('Namess.xlsx') 
    d.to_excel(GFG) 
    GFG.save()
    data=pd.read_excel('Namess.xlsx')
    data.columns=['a', 'b']
    #print(data['a'])
    #date
    da=data['a'][0].split("\t")[2]
    chumma=ws3.cell(row = 1, column =1).value
    ws2.cell(row = 1, column =chumma).value=da
    update=chumma+1
    ws3.cell(row=1,column=1).value=update

    print(da)
    #name
    name=[]
    for i in data['a']:
        name.append(i.split("\t")[0])

    print(name)


    i=2
    d=pd.read_excel('book1.xlsx')
    a=len(d)
    a=a+1
    print(a)

    #d.columns=['a','date']
    #d.rows=['b']
    #print(len(d.rows))
    while (i<=a):
        name1=ws2.cell(row=i,column=1).value
        #namee2=ws2.cell(row=2,column=1).value
        #name3=ws2.cell(row=3,column=1).value
        #print(type(name1))
        #print(name1)
        #print(type(name[i]))
        #print(name[i])
        temp=len(name)
        j=0
        rakita=0
        while(j<temp):
            if(name1==name[j]):
                ws2.cell(row = i, column =chumma).value ="P"
                flag=ws2.cell(row = i, column =chumma).value
                print(flag)
                rakita=1
            j=j+1
        if(rakita==0):
            ws2.cell(row = i, column =chumma).value ="A"
            
        print(name1)
        #print(name2)
        #print(name3)
        i=i+1


    wb2.save("book1.xlsx")


